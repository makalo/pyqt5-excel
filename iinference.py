import os
import sys
import time
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
from PyQt5.QtCore import Qt, QTimer, QCoreApplication, QSettings,pyqtSignal,QObject
from PyQt5.QtGui import QPixmap, QPainter
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QGridLayout, QSplashScreen
from PyQt5.QtWidgets import QVBoxLayout,QMessageBox
from PyQt5.QtGui import QTextCursor
import UI_lan
from ToolsPackage import splitThread
from openpyxl import load_workbook



class Stream(QObject):
    """Redirects console output to text widget."""
    newText = pyqtSignal(str)

    def write(self, text):
        QtWidgets.QApplication.processEvents()
        self.newText.emit(str(text))

class anaxcelhandler(QtWidgets.QMainWindow, UI_lan.Ui_MainWindow):

    def __init__(self, parent=None):
        super(anaxcelhandler, self).__init__(parent)
        if getattr(sys, 'frozen', False):
            self.frozen = 'ever so'
            self.bundle_dir = sys._MEIPASS
        else:
            self.bundle_dir = os.path.dirname(os.path.abspath(__file__))
        self.setupUi(self)
        self.setWindowIcon(QtGui.QIcon(self.bundle_dir + '/icons/icon.png'))
        self.setStyleSheet(open("Dark/darkstyle.qss", "r").read())
        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.pushButtonbrowse.clicked.connect(self.openFileNamesDialog)
        self.pushButtonclear.clicked.connect(self.clearwidget)
        self.pushButtonselall.clicked.connect(self.selectall)
        self.pushButtonload.clicked.connect(self.LoadProcess)
        self.pushButtonsplit.clicked.connect(self.SplitProcess)
        self.pushButtonmerge.clicked.connect(self.mergeProcess)
        self.pushButtonanalyse.clicked.connect(self.analyseProcess)
        self.pushButtonmakalo.clicked.connect(self.makaloProcess)

        self.statusbar.showMessage('兰神专属')
        self.comboBoxfiletype.addItems(['xlsx','xls'])

        #==========log=====
        sys.stdout = Stream(newText=self.onUpdateText)
        #==========log=====
    def onUpdateText(self, text):
        """Write console output to text widget."""
        cursor = self.textBrowserlog.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textBrowserlog.setTextCursor(cursor)
        self.textBrowserlog.ensureCursorVisible()


    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)
        elif self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)

    def clearwidget(self):
        self.listWidget.clear()
        self.tableWidget.clear()

    def selectall(self):
        self.listWidget.selectAll()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "关于对话框", '请先加载文件')

    def xlsProcess(self):
        self.tableWidget.clear()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "关于对话框", '请先选择文件')
        xlsfiles = []
        for i in list(items):
            xlsfiles.append(i.text())
        wkbk = xlwt.Workbook()
        outsheet = wkbk.add_sheet('Sheet1')
        outrow_idx = 0
        for f in xlsfiles:
            print('正在加载 {}'.format(os.path.split(f)[-1]))
            insheet = xlrd.open_workbook(f).sheets()[0]
            for row_idx in range(insheet.nrows):
                for col_idx in range(insheet.ncols):
                    outsheet.write(outrow_idx, col_idx, insheet.cell_value(row_idx, col_idx))
                outrow_idx += 1
        wkbk.save(r'combined.xls')
        # use on_demand=True to avoid loading worksheet data into memory
        book = xlrd.open_workbook("combined.xls", on_demand=True)
        sheet = book.sheet_by_index(0)
        num_rows = sheet.nrows
        num_col = sheet.ncols
        self.tableWidget.setRowCount(num_rows)
        self.tableWidget.setColumnCount(num_col)
        for col in range(num_col):
            for row in range(num_rows):
                cell = sheet.cell(row, col)
                if (not cell.value == "") and (not cell.value == " "):
                    self.tableWidget.setItem(row, col, QTableWidgetItem(str(cell.value)))
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()

    def xlsxprocess(self):
        self.tableWidget.clear()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "关于对话框", '请先选择文件')
        xlsfiles = []
        for i in list(items):
            xlsfiles.append(str(i.text()))

        wb = load_workbook(filename=xlsfiles[0])
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]
        num_row = ws.max_row
        num_column = ws.max_column
        self.tableWidget.setColumnCount(num_column)
        self.tableWidget.setRowCount(num_row)
        for i in range(1,num_row+1):
            for j in range(1,num_column+1):
                v = ws.cell(row=i, column=j).value
                self.tableWidget.setItem(i-1, j-1, QTableWidgetItem(str(v)))
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()

    def LoadProcess(self):
        if self.comboBoxfiletype.currentIndex() == 1:  # xls
            QMessageBox.about(self, "关于对话框", '不支持 xls 格式文件')
            # self.xlsProcess()
            # for colindex in range(self.tableWidget.columnCount()):
            #     self.comboBoxfiletypeX.addItem(str(colindex))
            #     self.comboBoxfiletypeY.addItem(str(colindex))

        elif self.comboBoxfiletype.currentIndex() == 0:  # xlsx
            self.xlsxprocess()
            for colindex in range(self.tableWidget.columnCount()):
                self.comboBoxfiletypeX.addItem(str(colindex))
                self.comboBoxfiletypeY.addItem(str(colindex))
        print('可以预览文件')

    def SplitProcess(self):
        x = self.comboBoxfiletypeX.currentIndex()
        y = self.comboBoxfiletypeY.currentIndex()
        print(x,y)
        if x == 0 or y == 0:
            QMessageBox.about(self, "关于对话框", '请先选择拆分关键词位置')
        if x == -1 or y == -1:
            QMessageBox.about(self, "关于对话框", '请先选择文件并load文件')
        else:
            items = self.listWidget.selectedItems()
            xlsfiles = []
            for i in list(items):
                xlsfiles.append(i.text())
            try:
                self.splitThread = splitThread(idx = [x,y],files = xlsfiles)
                self.splitThread.split_signal.connect(self.set_progressbar_value)
                self.splitThread.split_signal_lcd.connect(self.set_lcdnumber_value)
                self.splitThread.start()
            except:
                QMessageBox.about(self, "关于对话框", '拆分{}出现错误'.format(base_name))
    def mergeProcess(self):
        QMessageBox.about(self, "关于对话框", '此功能为付费功能')
    def analyseProcess(self):
        QMessageBox.about(self, "关于对话框", '此功能为付费功能')
    def makaloProcess(self):
        QMessageBox.question(self, "提问对话框", "感谢一下makalo吧？", QMessageBox.Yes | QMessageBox.No)

    def set_progressbar_value(self, value):
        self.progressBar.setValue(value)
    def set_lcdnumber_value(self,value):
        self.lcdNumber.display(value)


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = anaxcelhandler()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
