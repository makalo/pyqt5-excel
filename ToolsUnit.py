from openpyxl import Workbook, load_workbook,styles
from openpyxl.formula.translate import Translator
import os
import copy
import math
from openpyxl.utils import get_column_letter
from utils import get_key,get_merge_cell_list,get_merge_map,set_style,assign_style,idx2letter

def split_excel(path,base_info,signal = None):
    wb = load_workbook(filename=path)
    sheet_names = list(base_info.keys())
    valid_sheets = []
    invalid_sheets = []
    for s in sheet_names:
        base = base_info[s]
        if len(base) != 0 and base[0][0] != '':
            valid_sheets.append(s)
        else:
            invalid_sheets.append(s)

    dict_idx = get_key(wb,valid_sheets,base_info)


    wbs_split = []
    names_split = []

    num_keys = len(list(dict_idx.keys()))
    count = 0
    for k,dict_sheet in dict_idx.items():
        print('开始拆关键词:{}'.format(k))
        wb_tmp = Workbook()
        wb_tmp.remove(wb_tmp[wb_tmp.sheetnames[0]])

        for sheet,idxes in dict_sheet.items():
            ws = wb[sheet]
            num_row = ws.max_row
            num_column = ws.max_column

            _,rg = base_info[sheet]
            head_idx = list(range(1,rg[0]))
            tail_idx = list(range(num_row+1 if rg[1] == 'last' else rg[1]+1,num_row+1))

            # ws_rows = list(ws.values)
            ws_tmp = wb_tmp.create_sheet(sheet)
            idxes = head_idx+idxes+tail_idx

            #======合并单元格=======
            merge_idx = ws.merged_cells
            merge_idx = get_merge_cell_list(merge_idx)
            for m_idx in merge_idx:
                map_idx = get_merge_map(m_idx,idxes)
                if map_idx is not None:
                    ws_tmp.merge_cells(map_idx)
            #======合并单元格=======

            for i in range(1,len(idxes)+1):
                idx = idxes[i-1]
                # print(ws.row_dimensions[idx].height)
                # ws_tmp.append(ws_rows[idx-1])
                for j in range(1,num_column+1):
                    ws_tmp.row_dimensions[i].height = ws.row_dimensions[idx].height
                    ws_tmp.column_dimensions[get_column_letter(j)].width = ws.column_dimensions[get_column_letter(j)].width
                    cell = ws.cell(row=idx, column=j)
                    value = cell.value
                    #============公式=====
                    if isinstance(value,str) and '=' in value:
                        value = Translator(value, origin=idx2letter([idx,j])).translate_formula(idx2letter([i,j]))
                    #============公式=====
                    cell_tmp = ws_tmp.cell(row=i, column=j,value = value)
                    assign_style(cell_tmp,cell)
            # set_style(ws_tmp)
        for sheet in invalid_sheets:
            ws = wb[sheet]
            num_row = ws.max_row
            num_column = ws.max_column
            ws_tmp = wb_tmp.create_sheet(sheet)
            rows = ws.values
            for row in rows:
                ws_tmp.append(row)


        wbs_split.append(wb_tmp)
        names_split.append(k)
        count += 1
        if signal is not None:
            proess = count / num_keys * 100
            signal.emit(int(proess))
        wb.close()
    return wbs_split,names_split


if __name__ == "__main__":
    # base_info = {'省级产业集聚区地下水污染防治有关工作开展情况调度表':[[2,2],[3,'last']],
    #             '以化工等行业为主的产业集聚区':[]}
    base_info = {'其他':[],'数据字典':[],'Sheet1':[[3,2],[4,33]]}
    wbs_split,names_split = split_excel('20210223.xlsx',base_info)
    root = './debug'
    if not os.path.exists(root):
        os.makedirs(root)
    for wb,name in zip(wbs_split,names_split):
        print(name)
        path = os.path.join(root,name+'.xlsx')
        wb.save(path)









