import os
import sys
import time
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
from PyQt5.QtCore import Qt, QTimer, QCoreApplication, QSettings,pyqtSignal,QObject
from PyQt5.QtGui import QPixmap, QPainter
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QGridLayout, QSplashScreen
from PyQt5.QtWidgets import QVBoxLayout,QMessageBox
from PyQt5.QtGui import QTextCursor
import UI_lan
from ToolsPackage import splitThread
from openpyxl import load_workbook
from utils import get_column_letter,assign_style_qt,get_merge_cell_list
import webbrowser
import qdarkstyle
class Stream(QObject):
    """Redirects console output to text widget."""
    newText = pyqtSignal(str)

    def write(self, text):
        QtWidgets.QApplication.processEvents()
        self.newText.emit(str(text))

class anaxcelhandler(QtWidgets.QMainWindow, UI_lan.Ui_MainWindow):

    def __init__(self, parent=None):
        super(anaxcelhandler, self).__init__(parent)
        if getattr(sys, 'frozen', False):
            self.frozen = 'ever so'
            self.bundle_dir = sys._MEIPASS
        else:
            self.bundle_dir = os.path.dirname(os.path.abspath(__file__))
        self.setupUi(self)
        #self.setWindowIcon(QtGui.QIcon(self.bundle_dir + '/icons/icon.png'))
        #self.setStyleSheet(open("Dark/darkstyle.qss", "r").read())
        # self.setStyleSheet(open("qss/1.qss", "r").read())

        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.pushButtonbrowse.clicked.connect(self.openFileNamesDialog)
        self.pushButtonclear.clicked.connect(self.clearwidget)
        self.pushButtonselall.clicked.connect(self.selectall)
        self.pushButtonload.clicked.connect(self.LoadProcess)
        self.pushButtonsplit.clicked.connect(self.SplitProcess)
        self.pushButtonmerge.clicked.connect(self.mergeProcess)
        self.pushButtonanalyse.clicked.connect(self.analyseProcess)
        self.pushButtonmakalo.clicked.connect(self.makaloProcess)
        self.pushButton_link.clicked.connect(self.linkProcess)

        self.statusbar.showMessage('兰神专属')
        self.comboBoxfiletype.addItems(['xlsx','xls'])

        #==========log=====
        sys.stdout = Stream(newText=self.onUpdateText)
        #==========log=====

        #==========show====
        self.flag_confirm = False
        self.activate_file = [None,None]
        self.comboBox_wb.activated.connect(self.wbActivated)
        self.comboBox_ws.activated.connect(self.wsActivated)
        self.tableWidget.itemClicked.connect(self.handleItemClick)
        self.pushButton_clear_idx.clicked.connect(self.clear_idx)
        self.pushButton_confirm_idx.clicked.connect(self.confirm_idx)
        #==========show====

    def use_palette(self):
        self.setWindowTitle("设置背景图片")
        window_pale = QtGui.QPalette()
        window_pale.setBrush(self.backgroundRole(),   QtGui.QBrush(QtGui.QPixmap(self.bundle_dir + '/icons/bg.jpeg')))
        self.setPalette(window_pale)
    def onUpdateText(self, text):
        """Write console output to text widget."""
        cursor = self.textBrowserlog.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textBrowserlog.setTextCursor(cursor)
        self.textBrowserlog.ensureCursorVisible()


    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)
        elif self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)

    def clearwidget(self):
        self.listWidget.clear()
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()
    def clearcontext_all(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()
    def clearcontext_show(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
    def clear_idx(self):
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()
    def confirm_idx(self):
        x = self.comboBox_x.itemText(self.comboBox_x.currentIndex())
        y = self.comboBox_y.itemText(self.comboBox_y.currentIndex())

        r1 = self.comboBox_r1.itemText(self.comboBox_r1.currentIndex())
        r2 = self.comboBox_r2.itemText(self.comboBox_r2.currentIndex())

        wb = self.comboBox_wb.itemText(self.comboBox_wb.currentIndex())
        ws = self.comboBox_ws.itemText(self.comboBox_ws.currentIndex())

        if wb == '' or ws == '':
            QMessageBox.about(self, "hi,兰神", '先load文件')
        else:
            x = int(x) if x != '' else x
            y = int(y) if y != '' else y
            r1 = int(r1) if r1 != '' else r1
            r2 = int(r2) if r2 != '' else r2

            if self.checkBox_book.isChecked():
                print('book')
                key_idx = [x,y]
                rg = [r1,'last']
                for wb_k in self.infos.keys():
                    ws_keys = self.infos[wb_k]['sheet_names']
                    for ws_k in ws_keys.keys():
                        self.infos[wb_k]['sheet_names'][ws_k] = [key_idx,rg]
            elif self.checkBox_sheet.isChecked():
                print('sheet')
                key_idx = [x,y]
                rg = [r1,'last']
                ws_keys = self.infos[wb]['sheet_names']
                for ws_k in ws_keys.keys():
                    self.infos[wb]['sheet_names'][ws_k] = [key_idx,rg]
            else:
                print('cell')
                key_idx = [x,y]
                rg = [r1,r2]
                self.infos[wb]['sheet_names'][ws] = [key_idx,rg]
            self.flag_confirm = True


    def selectall(self):
        self.listWidget.selectAll()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi,兰神", '请先加载文件')

    def LoadProcess(self):
        self.clearcontext_all()
        if self.comboBoxfiletype.currentIndex() == 1:  # xls
            QMessageBox.about(self, "hi,兰神", '不支持 xls 格式文件')
        elif self.comboBoxfiletype.currentIndex() == 0:  # xlsx
            items = self.listWidget.selectedItems()
            if len(items) == 0:
                QMessageBox.about(self, "hi,兰神", '请先选择文件')
            else:
                self.infos = {}
                for i in list(items):
                    file_path = str(i.text())
                    wb = load_workbook(filename=file_path)
                    name = os.path.split(file_path)[-1]

                    sheet_names = wb.sheetnames

                    sheets_dict = {}
                    for s in sheet_names:
                        sheets_dict[s] = []
                    self.infos[name] = {'path':file_path,'sheet_names':sheets_dict}
                    wb.close()
                for k in self.infos.keys():
                    self.comboBox_wb.addItem(k)
                k = self.comboBox_wb.itemText(0)
                sheets = list(self.infos[k]['sheet_names'].keys())
                for s in sheets:
                    self.comboBox_ws.addItem(s)
                self.activate_file[0] = self.infos[k]['path']
                self.activate_file[1] = list(self.infos[k]['sheet_names'].keys())[0]

                self.show_excel()
        print('可以预览文件')
    def wbActivated(self,index):
        self.clearcontext_show()
        wb_k = self.comboBox_wb.itemText(index)
        sheets = list(self.infos[wb_k]['sheet_names'].keys())
        self.comboBox_ws.clear()
        for s in sheets:
            self.comboBox_ws.addItem(s)
        self.activate_file[0] = self.infos[wb_k]['path']
        self.activate_file[1] = list(self.infos[wb_k]['sheet_names'].keys())[0]
        self.show_excel()


    def wsActivated(self,index):
        ws_k = self.comboBox_ws.itemText(index)
        self.activate_file[1] = ws_k
        self.show_excel()

    def handleItemClick(self,item):
        cont = item.text()
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        row = item.row()+1
        column = item.column()+1
        #=======对合并的单元格取idx
        for p in self.merge_position:
            if row == p[0] and column == p[1]:
                row = row + (p[2]-p[0])
                break
        #=======对合并的单元格取idx
        self.comboBox_x.addItem(str(row))
        self.comboBox_y.addItem(str(column))
        self.comboBox_r1.addItem(str(row+1))


    def show_excel(self):
        self.merge_position = []
        path = self.activate_file[0]
        sheetname = self.activate_file[1]
        wb = load_workbook(filename=path)
        ws = wb[sheetname]
        num_row = ws.max_row
        num_column = ws.max_column
        self.tableWidget.setColumnCount(num_column)
        self.tableWidget.setRowCount(num_row)

        #======合并单元格=======
        merge_idx = ws.merged_cells
        merge_idx = get_merge_cell_list(merge_idx)

        for i in range(len(merge_idx)):
            m_idx = merge_idx[i]
            self.tableWidget.setSpan(m_idx[0]-1, m_idx[1]-1, m_idx[2]-m_idx[0]+1, m_idx[3]-m_idx[1]+1)
            self.merge_position.append([m_idx[0],m_idx[1],m_idx[2]])#[x1,y1,range]
        #======合并单元格=======

        #======单元格大小=======
        for i in range(1,num_row+1):
            h = ws.row_dimensions[i].height
            if h is not None:
                self.tableWidget.setRowHeight(i-1,h)
        # for i in range(1,num_column+1):
        #     w = ws.column_dimensions[get_column_letter(i)].width
        #     if w is not None:
        #         self.tableWidget.setColumnWidth(i-1,w)
        #======单元格大小=======

        self.comboBox_r2.clear()
        for i in range(1,num_row+1):
            self.comboBox_r2.addItem(str(num_row-i+1))
            row_sizes = []
            for j in range(1,num_column+1):
                cell = ws.cell(row=i, column=j)
                if cell.value is not None:
                    item = QTableWidgetItem(str(cell.value))
                    assign_style_qt(item,cell)
                else:
                    item = QTableWidgetItem()
                self.tableWidget.setItem(i-1, j-1, item)

        # self.tableWidget.resizeColumnsToContents()
        # self.tableWidget.resizeRowsToContents()
        wb.close()

    def SplitProcess(self):
        if not self.flag_confirm:
            QMessageBox.about(self, "hi,兰神", '请先选择文件并load文件,并选择拆分关键词')
        else:
            try:
                self.splitThread = splitThread(self.infos)
                self.splitThread.split_signal.connect(self.set_progressbar_value)
                self.splitThread.split_signal_lcd.connect(self.set_lcdnumber_value)
                self.splitThread.start()
            except:
                QMessageBox.about(self, "hi,兰神", '拆分{}出现错误'.format(base_name))
        self.flag_confirm = False
    def mergeProcess(self):
        QMessageBox.about(self, "hi,兰神", '此功能为付费功能')
    def analyseProcess(self):
        QMessageBox.about(self, "hi,兰神", '此功能为付费功能')
    def makaloProcess(self):
        webbrowser.open('https://blog.csdn.net/guicai1647855685?spm=1010.2135.3001.5421')
    def linkProcess(self):
        webbrowser.open('https://github.com/makalo')
    def set_progressbar_value(self, value):
        self.progressBar.setValue(value)
    def set_lcdnumber_value(self,value):
        self.lcdNumber.display(value)



app = QtWidgets.QApplication(sys.argv)
window = anaxcelhandler()
# setup stylesheet
app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
window.show()
sys.exit(app.exec_())

