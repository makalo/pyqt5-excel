from openpyxl import Workbook, load_workbook,styles
import os
import copy
import math
from openpyxl.utils import get_column_letter, column_index_from_string
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt
def get_key(wb,valid_sheets,base_info):
    sheet_names = valid_sheets
    # print(sheet_names)
    dict_idx = {}
    keys = []
    for name_s in sheet_names:
        ws = wb[name_s]
        idx,rg = base_info[name_s]
        _,y = idx
        num_row = ws.max_row
        num_column = ws.max_column

        merge_idx = ws.merged_cells
        merge_idx = get_merge_cell_list(merge_idx)

        for i in range(rg[0],num_row+1 if rg[1] == 'last' else rg[1]+1):
            k = ws.cell(row=i, column=y).value
            if k is None:
                for m_idx in merge_idx:
                    if y == m_idx[1] and i >= m_idx[0] and i<= m_idx[2]:
                        k = keys[-1]
                        break
                    else:
                        continue
                if k is None:
                    continue
            keys.append(k)
            if not k in dict_idx.keys():
                dict_idx[k] = {name_s:[i]}
            elif not name_s in dict_idx[k].keys():
                dict_idx[k][name_s] = [i]
            else:
                l = dict_idx[k][name_s]
                l.append(i)
                dict_idx[k][name_s] = l
    return dict_idx

def get_merge_cell_list(merge_idx):
    merge_idx = list(merge_idx)
    merge_list = []
    for i in range(len(merge_idx)):
        merge = merge_idx[i]
        row_min, row_max, col_min, col_max = merge.min_row, merge.max_row, merge.min_col, merge.max_col
        # merge_list.append([row_min, row_max, col_min, col_max])
        merge_list.append([row_min, col_min, row_max, col_max])
    return merge_list
def get_merge_map(merge_idx,idx):
    row_min, col_min,row_max , col_max = merge_idx
    col_min = get_column_letter(col_min)
    col_max = get_column_letter(col_max)
    try:
        x1 = idx.index(row_min) + 1
        x2 = idx.index(row_max) + 1
        s = '{}{}:{}{}'.format(col_min,x1,col_max,x2)
        return s
    except:
        return None

def idx2letter(idx):
    x,y = idx
    y = get_column_letter(y)
    s = '{}{}'.format(y,x)
    return s


def set_style(ws):
    align = styles.Alignment(horizontal='center',vertical='center')
    font_18 = styles.Font(size=18)

    num_row = ws.max_row #最大行数
    num_column = ws.max_column #最大列数
    for i in range(1,num_row+1):
        for j in range(1,num_column+1):
            cell = ws.cell(row=i, column=j)
            cell.alignment = align
            cell.font = font_18
def assign_style(target_cell,source_cell):
    target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)

def color(value):
    digit = list(map(str, range(10))) + list("ABCDEF")
    if isinstance(value, tuple):
        string = '#'
        for i in value:
            a1 = i // 16
            a2 = i % 16
            string += digit[a1] + digit[a2]
        return string
    elif isinstance(value, str):
        a1 = digit.index(value[1]) * 16 + digit.index(value[2])
        a2 = digit.index(value[3]) * 16 + digit.index(value[4])
        a3 = digit.index(value[5]) * 16 + digit.index(value[6])
        return (a1, a2, a3)
    else:
        return (0, 0, 0)
def assign_style_qt(target_cell,source_cell):
    #字体，大小，颜色，加粗
    font = QFont()   #实例化字体对象
    font.setFamily(source_cell.font.name)  #字体
    font.setBold(source_cell.font.bold)  #加粗
    font.setPointSize(source_cell.font.size)   #字体大小
    target_cell.setFont(font)
    #居中
    target_cell.setTextAlignment(Qt.AlignCenter | Qt.AlignCenter)
    #背景
    # print(source_cell.fill.bgColor.rgb)
    # target_cell.setBackground(QtGui.QColor(100,100,150))










